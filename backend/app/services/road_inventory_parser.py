"""
Parser for Caltrans TSN Highway Inventory Excel exports.

Expected format: single sheet named "Sheet 1", row 1 = headers (THY_* columns),
rows 2+ = data.  74 columns as of the 2025-06-08 extract.

Confirmed column names from CA HIGHWAYS 06.08.2025.xlsx (74 columns, 59 765 data rows).
THY_COUNTY_CODE stores 3-letter codes ('ORA', 'SAC', etc.).
THY_DISTRICT_CODE stores string district numbers ('01'–'12').
"""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

import openpyxl

# ---------------------------------------------------------------------------
# Column map: Excel header → normalized field name
# Only columns that land in named road_segments columns are listed here.
# Every other THY_* column is folded into raw_json.
# ---------------------------------------------------------------------------
_NORMALIZED_COLS: dict[str, str] = {
    "THY_ID":                       "thy_id",
    "THY_DISTRICT_CODE":            "district_code",
    "THY_COUNTY_CODE":              "county_code",
    "THY_ROUTE_NAME":               "route_name",
    "THY_ROUTE_SUFFIX_CODE":        "route_suffix_code",
    "THY_PM_PREFIX_CODE":           "pm_prefix_code",
    "THY_BEGIN_PM_AMT":             "begin_pm",
    "THY_END_PM_AMT":               "end_pm",
    "THY_PM_SUFFIX_CODE":           "pm_suffix_code",
    "THY_LENGTH_MILES_AMT":         "length_miles",
    "THY_LT_SURF_TYPE_CODE":        "left_surface_type",
    "THY_LT_LANES_AMT":             "left_lanes",
    "THY_LT_O_SHD_TOT_WIDTH_AMT":  "left_shoulder_width",
    "THY_RT_SURF_TYPE_CODE":        "right_surface_type",
    "THY_RT_LANES_AMT":             "right_lanes",
    "THY_RT_O_SHD_TOT_WIDTH_AMT":  "right_shoulder_width",
    "THY_MEDIAN_TYPE_CODE":         "median_type",
    "THY_MEDIAN_WIDTH_AMT":         "median_width",
    "THY_HIGHWAY_ACCESS_CODE":      "access_code",
    "THY_TERRAIN_CODE":             "terrain_code",
    "THY_DESIGN_SPEED_AMT":         "design_speed",
    "THY_ADT_AMT":                  "adt",
    "THY_LANDMARK_SHORT_DESC":      "landmark_short_desc",
    "THY_FUNCTIONAL_CLASS_CODE":    "functional_class_code",
    "THY_MAINT_SVC_LVL_CODE":       "maintenance_service_level_code",
    "THY_FEDERAL_AID_CODE":         "federal_aid_code",
    "THY_SCENIC_FREEWAY_CODE":      "scenic_freeway_code",
    "THY_EXTRACT_DATE":             "extract_date",
}

REQUIRED_COLUMNS: frozenset[str] = frozenset({
    "THY_COUNTY_CODE",
    "THY_ROUTE_NAME",
    "THY_BEGIN_PM_AMT",
    "THY_END_PM_AMT",
})


@dataclass
class ParseResult:
    rows: list[dict[str, Any]]
    row_count: int
    skipped_count: int
    warnings: list[str]
    extract_date: date | None = None


class ParseError(ValueError):
    """Raised for file-level errors (wrong format, missing required columns)."""


def parse_excel(file_bytes: bytes) -> ParseResult:
    """Parse a Caltrans TSN highway inventory Excel file.

    Returns a ParseResult with normalized row dicts ready for DB insert.
    Row-level errors are counted as skipped; they do not raise.
    File-level errors (bad format, missing columns) raise ParseError.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError(f"Cannot open workbook: {exc}") from exc

    ws = wb.active
    if ws is None:
        raise ParseError("Workbook has no active sheet")

    raw_rows = ws.iter_rows(values_only=True)

    try:
        header_row = next(raw_rows)
    except StopIteration:
        raise ParseError("Workbook sheet is empty")
    finally:
        wb.close()

    headers = [str(h).strip() if h is not None else "" for h in header_row]

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ParseError(
            f"Missing required columns: {sorted(missing)}. "
            f"Found {len(headers)} columns in header row."
        )

    # Reopen for data rows (read_only iterators can't be rewound)
    wb2 = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws2 = wb2.active
    data_iter = ws2.iter_rows(min_row=2, values_only=True)

    rows: list[dict[str, Any]] = []
    skipped = 0
    warnings: list[str] = []
    extract_date: date | None = None

    for row_num, raw in enumerate(data_iter, start=2):
        raw_dict = dict(zip(headers, raw))
        result = _normalize_row(raw_dict, row_num, warnings)
        if result is None:
            skipped += 1
            continue
        rows.append(result)
        if extract_date is None and result.get("extract_date"):
            extract_date = result["extract_date"]

    wb2.close()

    if skipped > 0:
        warnings.append(f"Skipped {skipped} rows due to validation errors (see above).")

    return ParseResult(
        rows=rows,
        row_count=len(rows),
        skipped_count=skipped,
        warnings=warnings,
        extract_date=extract_date,
    )


def _normalize_row(
    raw: dict[str, Any],
    row_num: int,
    warnings: list[str],
) -> dict[str, Any] | None:
    """Normalize one raw row dict.  Returns None to skip the row."""

    # Required fields
    county_code = _str(raw.get("THY_COUNTY_CODE"))
    route_name = _str(raw.get("THY_ROUTE_NAME"))
    begin_pm = _decimal(raw.get("THY_BEGIN_PM_AMT"))
    end_pm = _decimal(raw.get("THY_END_PM_AMT"))

    if not county_code:
        warnings.append(f"Row {row_num}: blank THY_COUNTY_CODE — skipped")
        return None
    if not route_name:
        warnings.append(f"Row {row_num}: blank THY_ROUTE_NAME — skipped")
        return None
    if begin_pm is None:
        warnings.append(f"Row {row_num}: invalid THY_BEGIN_PM_AMT={raw.get('THY_BEGIN_PM_AMT')!r} — skipped")
        return None
    if end_pm is None:
        warnings.append(f"Row {row_num}: invalid THY_END_PM_AMT={raw.get('THY_END_PM_AMT')!r} — skipped")
        return None
    if begin_pm >= end_pm:
        warnings.append(f"Row {row_num}: begin_pm={begin_pm} >= end_pm={end_pm} — skipped")
        return None

    normalized: dict[str, Any] = {
        "thy_id":           _int(raw.get("THY_ID")),
        "district_code":    _str(raw.get("THY_DISTRICT_CODE")),
        "county_code":      county_code.strip().upper(),
        "route_name":       normalize_route(route_name) or route_name,
        "route_suffix_code": _str(raw.get("THY_ROUTE_SUFFIX_CODE")),
        "pm_prefix_code":   _str(raw.get("THY_PM_PREFIX_CODE")),
        "begin_pm":         begin_pm,
        "end_pm":           end_pm,
        "pm_suffix_code":   _str(raw.get("THY_PM_SUFFIX_CODE")),
        "length_miles":     _decimal(raw.get("THY_LENGTH_MILES_AMT")),
        "left_surface_type":  _str(raw.get("THY_LT_SURF_TYPE_CODE")),
        "left_lanes":         _int(raw.get("THY_LT_LANES_AMT")),
        "left_shoulder_width": _decimal(raw.get("THY_LT_O_SHD_TOT_WIDTH_AMT")),
        "right_surface_type": _str(raw.get("THY_RT_SURF_TYPE_CODE")),
        "right_lanes":        _int(raw.get("THY_RT_LANES_AMT")),
        "right_shoulder_width": _decimal(raw.get("THY_RT_O_SHD_TOT_WIDTH_AMT")),
        "median_type":      _str(raw.get("THY_MEDIAN_TYPE_CODE")),
        "median_width":     _decimal(raw.get("THY_MEDIAN_WIDTH_AMT")),
        "access_code":      _str(raw.get("THY_HIGHWAY_ACCESS_CODE")),
        "terrain_code":     _str(raw.get("THY_TERRAIN_CODE")),
        "design_speed":     _int(raw.get("THY_DESIGN_SPEED_AMT")),
        "adt":              _int(raw.get("THY_ADT_AMT")),
        "landmark_short_desc": _str(raw.get("THY_LANDMARK_SHORT_DESC")),
        "functional_class_code": _str(raw.get("THY_FUNCTIONAL_CLASS_CODE")),
        "maintenance_service_level_code": _str(raw.get("THY_MAINT_SVC_LVL_CODE")),
        "federal_aid_code": _str(raw.get("THY_FEDERAL_AID_CODE")),
        "scenic_freeway_code": _str(raw.get("THY_SCENIC_FREEWAY_CODE")),
        "extract_date":     _date(raw.get("THY_EXTRACT_DATE")),
    }

    # Everything not in _NORMALIZED_COLS goes into raw_json
    raw_extras: dict[str, Any] = {}
    for col, val in raw.items():
        if col not in _NORMALIZED_COLS:
            raw_extras[col] = _json_safe(val)
    normalized["raw_json"] = json.dumps(raw_extras) if raw_extras else None

    return normalized


# ---------------------------------------------------------------------------
# Type coercers — all return None on failure (never raise)
# ---------------------------------------------------------------------------

def _str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _int(v: Any) -> int | None:
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _decimal(v: Any) -> Decimal | None:
    if v is None:
        return None
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return None


def _date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _json_safe(v: Any) -> Any:
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    return v


# ---------------------------------------------------------------------------
# Route normalization — used by both parser and lookup
# ---------------------------------------------------------------------------

def normalize_route(route: str | None) -> str | None:
    """Strip leading zeros and whitespace for consistent matching.

    '001' → '1', 'US 101' → '101', '  101  ' → '101'
    """
    if not route:
        return None
    r = route.strip().upper()
    # Strip common prefix words (US, SR, CA, HWY) sometimes found in user input
    for prefix in ("US ", "SR ", "CA ", "HWY ", "HIGHWAY "):
        if r.startswith(prefix):
            r = r[len(prefix):]
            break
    # Remove leading zeros
    r = r.lstrip("0") or "0"
    return r
